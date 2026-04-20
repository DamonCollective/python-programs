#!/usr/bin/env python3
"""
Search damoncollective@gmail.com for Piraeus Bank transaction emails
and extract payments to Μπακαρης Αγαθοκλης.

Piraeus masks beneficiary names: ΜΠΑΚΑΡΗΣ ΑΓΑΘΟΚΛΗΣ → ΜΠ*****Σ ΑΓ******Σ
We match by the ΜΠ...ΑΓ pattern in the Κύριος Δικαιούχος field.

Output: one Excel with two sheets — 2024 and 2025/2026.
"""

import os
import re
import base64

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
TOKEN_FILE  = os.path.join(SCRIPT_DIR, "token_damoncollective.json")
SCOPES      = ["https://www.googleapis.com/auth/gmail.modify"]
OUTPUT_FILE = os.path.expanduser("~/Desktop/bakaros_payments.xlsx")

SEARCH_QUERY = (
    'from:IBankV2Email@piraeusbank.gr '
    'subject:"Ενημέρωση Εγχρήματης Συναλλαγής" '
    'after:2023/12/31'
)

def get_service():
    creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
    return build("gmail", "v1", credentials=creds)

def get_body(payload):
    body = ""
    if "parts" in payload:
        for part in payload["parts"]:
            body += get_body(part)
    else:
        data = payload.get("body", {}).get("data", "")
        if data:
            body += base64.urlsafe_b64decode(data + "==").decode("utf-8", errors="replace")
    return body

def clean_html(body):
    text = re.sub(r'<style[^>]*>.*?</style>', '', body, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = re.sub(r'&nbsp;', ' ', text)
    text = re.sub(r'&amp;', '&', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def is_bakaros(text):
    text_lower = text.lower()
    if any(k in text_lower for k in ["μπακαρ", "bakaros", "αγαθοκλ"]):
        return True
    m = re.search(
        r'(?:Κύριος\s+Δικαιούχος|Δικαιούχος)\s*:\s*(ΜΠ[\*\w]*\s+ΑΓ[\*\w]*)',
        text, re.IGNORECASE
    )
    return bool(m)

def parse_email(text, email_date):
    def field(pattern):
        m = re.search(pattern, text, re.IGNORECASE)
        return m.group(1).strip() if m else ""

    # Extract masked name cleanly (first word = ΜΠ..., second = ΑΓ...)
    name = "ΜΠΑΚΑΡΗΣ ΑΓΑΘΟΚΛΗΣ"
    m = re.search(
        r'(?:Κύριος\s+Δικαιούχος|Δικαιούχος)\s*:\s*(ΜΠ[\*\w]*\s+ΑΓ[\*\w]*)',
        text, re.IGNORECASE
    )
    if m:
        name = m.group(1).strip()

    # Reason: grab everything between Αιτιολογία: and the next known field
    reason = field(r'Αιτιολογία\s*:\s*(.+?)(?=\s+Επιβάρυνση|\s+Εκτέλεση:|\s+supportebanking|\s+Email Επικοινωνίας|$)')
    if not reason:
        reason = field(r'Αιτιολογία\s*:\s*(.+)')

    amount = field(r'Ποσό Συναλλαγής\s*:\s*([\d.,]+(?:\s*EUR)?)')

    return {
        "date":               field(r'Ημερομηνία Εκτέλεσης\s*:\s*([\d/]+)') or email_date,
        "name":               name,
        "transaction_number": field(r'Κωδικός Συναλλαγής\s*:\s*([A-Z0-9]+)'),
        "amount":             amount,
        "reason":             reason,
    }

def date_sort_key(txn):
    d = txn["date"]  # DD/MM/YYYY
    try:
        return (d[6:10], d[3:5], d[0:2])
    except Exception:
        return ("0000", "00", "00")

def amount_to_float(s):
    s = re.sub(r'[^\d,.]', '', s).replace(',', '.')
    try:
        return float(s)
    except Exception:
        return 0.0

def build_sheet(ws, transactions, title):
    # Header
    header_row = ["Ημερομηνία", "Δικαιούχος", "Κωδικός Συναλλαγής", "Ποσό (EUR)", "Αιτιολογία"]
    ws.append(header_row)

    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 70

    thin      = Side(border_style="thin", color="AAAAAA")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    even_fill = PatternFill("solid", fgColor="EBF3FB")

    total_eur = 0.0
    for idx, txn in enumerate(sorted(transactions, key=date_sort_key), start=2):
        amt = amount_to_float(txn["amount"])
        total_eur += amt
        ws.append([
            txn["date"],
            txn["name"],
            txn["transaction_number"],
            amt,
            txn["reason"],
        ])
        for col, cell in enumerate(ws[idx], start=1):
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 5))
            if idx % 2 == 0:
                cell.fill = even_fill
            if col == 4:
                cell.number_format = '#,##0.00 "EUR"'
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Total row
    total_row_idx = ws.max_row + 1
    ws.cell(row=total_row_idx, column=3, value="ΣΥΝΟΛΟ")
    ws.cell(row=total_row_idx, column=3).font = Font(bold=True)
    ws.cell(row=total_row_idx, column=3).alignment = Alignment(horizontal="right")
    tc = ws.cell(row=total_row_idx, column=4, value=total_eur)
    tc.number_format = '#,##0.00 "EUR"'
    tc.font = Font(bold=True)
    tc.fill = PatternFill("solid", fgColor="D6E4F0")
    tc.alignment = Alignment(horizontal="right", vertical="center")

    ws.freeze_panes = "A2"
    return total_eur

def main():
    print("Authenticating...")
    service = get_service()

    print("Fetching all Piraeus transaction emails from 2024 onwards...")
    all_messages = []
    page_token = None
    while True:
        params = {"userId": "me", "q": SEARCH_QUERY, "maxResults": 500}
        if page_token:
            params["pageToken"] = page_token
        response = service.users().messages().list(**params).execute()
        all_messages.extend(response.get("messages", []))
        page_token = response.get("nextPageToken")
        if not page_token:
            break

    print(f"Found {len(all_messages)} transaction emails. Scanning for Bakaros...")

    txns_2024 = []
    txns_2025 = []

    for i, msg_ref in enumerate(all_messages):
        msg = service.users().messages().get(
            userId="me", id=msg_ref["id"], format="full"
        ).execute()

        headers  = {h["name"]: h["value"] for h in msg["payload"]["headers"]}
        date_str = headers.get("Date", "")

        try:
            from email.utils import parsedate_to_datetime
            dt         = parsedate_to_datetime(date_str)
            email_date = dt.strftime("%d/%m/%Y")
            year       = dt.year
            if year < 2024:
                continue
        except Exception:
            email_date = date_str[:10]
            year = 0

        body = get_body(msg["payload"])
        text = clean_html(body)

        if is_bakaros(text):
            txn = parse_email(text, email_date)
            print(f"  MATCH {year}: {txn['date']} | {txn['amount']} | {txn['transaction_number']} | {txn['reason'][:60]}")
            if year == 2024:
                txns_2024.append(txn)
            else:
                txns_2025.append(txn)

        elif (i + 1) % 50 == 0:
            print(f"  Checked {i+1}/{len(all_messages)}...")

    print(f"\n2024: {len(txns_2024)} payments | 2025+: {len(txns_2025)} payments")

    # ── Excel with two sheets ─────────────────────────────────────────────────

    wb = openpyxl.Workbook()

    ws2024 = wb.active
    ws2024.title = "2024"
    t2024 = build_sheet(ws2024, txns_2024, "2024")

    ws2025 = wb.create_sheet(title="2025-2026")
    t2025 = build_sheet(ws2025, txns_2025, "2025-2026")

    wb.save(OUTPUT_FILE)

    print(f"\nExcel saved: {OUTPUT_FILE}")
    print(f"  2024 total:      €{t2024:,.2f} ({len(txns_2024)} payments)")
    print(f"  2025/2026 total: €{t2025:,.2f} ({len(txns_2025)} payments)")
    print(f"  Grand total:     €{t2024 + t2025:,.2f}")

if __name__ == "__main__":
    main()
