#!/usr/bin/env python3
"""
Search damoncollective@gmail.com for Piraeus Bank transaction emails
and extract payments to Κωνσταντίνος Παπαπαναγιώτου — February & March 2026.

Piraeus masks beneficiary names. We match broadly on ΚΩ/παπαν/papanagiotou
in the Κύριος Δικαιούχος field or anywhere in the text.

Output: Excel with Date, Amount, Description columns.
"""

import os
import re
import base64

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
TOKEN_FILE  = os.path.join(SCRIPT_DIR, "token_damoncollective.json")
SCOPES      = ["https://www.googleapis.com/auth/gmail.modify"]
OUTPUT_FILE = os.path.expanduser("~/Documents/papapanagiotou_payments_feb_mar_2026.xlsx")

SEARCH_QUERY = (
    'from:IBankV2Email@piraeusbank.gr '
    'subject:"Ενημέρωση Εγχρήματης Συναλλαγής" '
    'after:2026/01/31 before:2026/04/01'
)


def get_service():
    creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
    return build("gmail", "v1", credentials=creds)


def get_body(payload):
    body = ""
    if "parts" in payload:
        for part in payload["parts"]:
            body += get_body(part)
    else:
        data = payload.get("body", {}).get("data", "")
        if data:
            body += base64.urlsafe_b64decode(data + "==").decode("utf-8", errors="replace")
    return body


def clean_html(body):
    text = re.sub(r'<style[^>]*>.*?</style>', '', body, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = re.sub(r'&nbsp;', ' ', text)
    text = re.sub(r'&amp;', '&', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def is_papapanagiotou(text):
    text_lower = text.lower()
    # Match plain name
    if any(k in text_lower for k in ["παπαπαναγ", "παπαναγ", "papapanag", "παπαπαν", "κωνσταντινοσ παπ"]):
        return True
    # Match Piraeus masked pattern: ΚΟ****Σ ΠΑ****Σ
    if re.search(r'ΚΟ[\*\w]+Σ\s+ΠΑ[\*\w]+Σ', text):
        return True
    return False


def parse_email(text, email_date):
    def field(pattern):
        m = re.search(pattern, text, re.IGNORECASE)
        return m.group(1).strip() if m else ""

    reason = field(r'Αιτιολογία\s*:\s*(.+?)(?=\s+Επιβάρυνση|\s+Εκτέλεση:|\s+supportebanking|\s+Email Επικοινωνίας|$)')
    if not reason:
        reason = field(r'Αιτιολογία\s*:\s*(.+)')

    amount = field(r'Ποσό Συναλλαγής\s*:\s*([\d.,]+(?:\s*EUR)?)')
    date   = field(r'Ημερομηνία Εκτέλεσης\s*:\s*([\d/]+)') or email_date

    return {"date": date, "amount": amount, "reason": reason}


def date_sort_key(txn):
    d = txn["date"]  # DD/MM/YYYY
    try:
        return (d[6:10], d[3:5], d[0:2])
    except Exception:
        return ("0000", "00", "00")


def amount_to_float(s):
    s = re.sub(r'[^\d,.]', '', s).replace(',', '.')
    try:
        return float(s)
    except Exception:
        return 0.0


def build_excel(transactions):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Φεβ-Μαρ 2026"

    headers = ["Ημερομηνία", "Ποσό (EUR)", "Αιτιολογία"]
    ws.append(headers)

    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 80

    thin      = Side(border_style="thin", color="AAAAAA")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    even_fill = PatternFill("solid", fgColor="EBF3FB")

    total = 0.0
    for idx, txn in enumerate(sorted(transactions, key=date_sort_key), start=2):
        amt = amount_to_float(txn["amount"])
        total += amt
        ws.append([txn["date"], amt, txn["reason"]])
        for col, cell in enumerate(ws[idx], start=1):
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 3))
            if idx % 2 == 0:
                cell.fill = even_fill
            if col == 2:
                cell.number_format = '#,##0.00 "EUR"'
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Total row
    tr = ws.max_row + 1
    ws.cell(row=tr, column=1, value="ΣΥΝΟΛΟ").font = Font(bold=True)
    tc = ws.cell(row=tr, column=2, value=total)
    tc.number_format = '#,##0.00 "EUR"'
    tc.font = Font(bold=True)
    tc.fill = PatternFill("solid", fgColor="D6E4F0")
    tc.alignment = Alignment(horizontal="right", vertical="center")

    ws.freeze_panes = "A2"
    wb.save(OUTPUT_FILE)
    return total


def main():
    print("Authenticating...")
    service = get_service()

    print("Fetching Piraeus transaction emails Feb–Mar 2026...")
    all_messages = []
    page_token = None
    while True:
        params = {"userId": "me", "q": SEARCH_QUERY, "maxResults": 500}
        if page_token:
            params["pageToken"] = page_token
        response = service.users().messages().list(**params).execute()
        all_messages.extend(response.get("messages", []))
        page_token = response.get("nextPageToken")
        if not page_token:
            break

    print(f"Found {len(all_messages)} transaction emails. Scanning for Παπαπαναγιώτου...")

    transactions = []
    for i, msg_ref in enumerate(all_messages):
        msg = service.users().messages().get(
            userId="me", id=msg_ref["id"], format="full"
        ).execute()

        headers  = {h["name"]: h["value"] for h in msg["payload"]["headers"]}
        date_str = headers.get("Date", "")

        try:
            from email.utils import parsedate_to_datetime
            dt         = parsedate_to_datetime(date_str)
            email_date = dt.strftime("%d/%m/%Y")
        except Exception:
            email_date = date_str[:10]

        body = get_body(msg["payload"])
        text = clean_html(body)

        if is_papapanagiotou(text):
            txn = parse_email(text, email_date)
            print(f"  MATCH: {txn['date']} | {txn['amount']} | {txn['reason'][:70]}")
            transactions.append(txn)
        elif (i + 1) % 20 == 0:
            print(f"  Checked {i+1}/{len(all_messages)}...")

    print(f"\nTotal matches: {len(transactions)}")

    if not transactions:
        print("No payments found for Παπαπαναγιώτου in Feb–Mar 2026.")
        return

    total = build_excel(transactions)
    print(f"Excel saved: {OUTPUT_FILE}")
    print(f"Total paid: €{total:,.2f} ({len(transactions)} payments)")


if __name__ == "__main__":
    main()
